from __future__ import annotations
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse
import hashlib
import json
import os
import re
import sys
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, timezone
from email.utils import parsedate_to_datetime
from html import escape
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urljoin
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape as xml_escape

import requests
from bs4 import BeautifulSoup

try:
    import win32com.client  # type: ignore
except Exception:
    win32com = None  # type: ignore


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
REPORT_DIR = BASE_DIR / "reports"
DEFAULT_CONFIG = BASE_DIR / "config.json"
DEFAULT_STATE = BASE_DIR / "state.json"
DEFAULT_RUN_SUMMARY = BASE_DIR / "run_summary.json"
KST = timezone(timedelta(hours=9))

SESSION = requests.Session()
SESSION.headers.update(
    {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        )
    }
)

DEFAULT_CONFIG_DATA: Dict[str, Any] = {
    "mail": {
        "to": "",
        "cc": "",
        "subject_prefix": "[부울경 행사 특별소통 알림]",
        "mode": "display"
    },
    "geocoding": {
        "enabled": True,
        "provider": "tmap",
        "region_hint": "부산 울산 경남"
    },
    "output": {
        "max_items_total": 30,
        "max_items_per_source": 10,
        "open_after_run": True
    },
    "filters": {
        "require_keyword": True,
        "require_region": True,
        "skip_expired_events": True,
        "region_keywords": [
            "부산", "울산", "경남", "경상남도", "김해", "창원", "양산",
            "진주", "거제", "통영", "사천", "밀양", "하동", "함안",
            "거창", "고성", "남해", "합천", "산청", "의령", "창녕"
        ],
        "include_keywords": [
            "행사", "축제", "공연", "페스티벌", "콘서트", "초대가수",
            "모집", "체험", "전시", "개최", "박람회", "문화", "집회",
            "걷기", "마라톤", "불꽃", "개막"
        ],
        "exclude_keywords": [
            "인사", "입찰", "채용", "분양", "공고문", "선거", "정당", "성료",
            "교육생", "직업교육훈련", "창업경진대회", "공공데이터", "수업 공개", "수업혁신",
            "헌혈", "장기등 기증", "센터 개소", "봉사단", "홍보관"
        ],
        "rss": {
            "fetch_article_detail": True,
            "max_article_fetch_per_source": 15,
            "article_timeout_seconds": 8,
            "article_max_chars": 6000,
            "min_signal_score": 4,
            "strong_keywords": [
                "축제", "공연", "콘서트", "페스티벌", "박람회", "엑스포", "전시",
                "에어쇼", "마라톤", "걷기", "대회", "체육대회", "로드쇼", "퍼레이드",
                "야행", "개막", "집회", "시위", "기자회견", "행진", "규탄", "총궐기"
            ],
            "weak_keywords": [
                "행사", "개최", "열다", "열린", "운영", "안내", "알림", "진행", "추진",
                "홍보", "캠페인", "교육", "모집", "포상", "표창", "봉사", "나눔", "방문",
                "점검", "취업지원", "공개모집", "서류심사", "기념식", "주간", "위촉",
                "협약", "심사", "발대식", "해단식", "실천활동", "소통", "공감"
            ],
            "venue_keywords": [
                "공원", "광장", "체육관", "경기장", "컨벤션", "전시장", "아트홀",
                "문화회관", "시청", "구청", "군청", "박물관", "대학", "대운동장",
                "역광장", "해변", "부두", "시장", "스타디움"
            ],
            "protest_keywords": [
                "집회", "시위", "기자회견", "행진", "규탄", "총궐기", "반대집회", "촛불"
            ]
        }
    },
    "ai_extraction": {
        "enabled": False,
        "provider": "openai",
        "model": "gpt-4o-mini",
        "only_for_rss": True,
        "fallback_on_missing": True,
        "min_body_chars": 200,
        "drop_non_events": False
    },
    "rules": {
        "default_network": "검토",
        "unknown_crowd_grade": "검토",
        "grade_thresholds": {
            "상": 10000,
            "중": 3000,
            "하": 0
        }
    },
    "sources": [
        {
            "name": "부산시 보도자료",
            "type": "busan_notice",
            "url": "https://www.busan.go.kr/nbnews"
        },
        {
            "name": "경남축제포털",
            "type": "gyeongnam_festa",
            "url": "https://festa.gyeongnam.go.kr/"
        },
        {
            "name": "부산 행사 뉴스",
            "type": "rss",
            "url": "https://news.google.com/rss/search?q=%EB%B6%80%EC%82%B0%20%ED%96%89%EC%82%AC%20when:1d&hl=ko&gl=KR&ceid=KR:ko"
        },
        {
            "name": "울산 행사 뉴스",
            "type": "rss",
            "url": "https://news.google.com/rss/search?q=%EC%9A%B8%EC%82%B0%20%ED%96%89%EC%82%AC%20when:1d&hl=ko&gl=KR&ceid=KR:ko"
        },
        {
            "name": "경남 행사 뉴스",
            "type": "rss",
            "url": "https://news.google.com/rss/search?q=%EA%B2%BD%EB%82%A8%20%ED%96%89%EC%82%AC%20when:1d&hl=ko&gl=KR&ceid=KR:ko"
        }
    ]
}

EVENT_HEADERS = [
    "구분",
    "기간",
    "Event명",
    "장소",
    "Type",
    "시작일",
    "종료일",
    "이벤트 등급",
    "이벤트 특성 구분",
    "공동망 여부",
    "예상운집인원(Peak Time)",
    "출처",
    "게시일",
    "원문링크",
    "요약",
    "수집일",
    "추출방식",
    "상태/비고",
]

WEEKDAYS = ["월", "화", "수", "목", "금", "토", "일"]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    if "<" in text and ">" in text:
        text = BeautifulSoup(text, "html.parser").get_text(" ", strip=True)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def ensure_runtime_files() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    if not DEFAULT_CONFIG.exists():
        write_json(DEFAULT_CONFIG, DEFAULT_CONFIG_DATA)
        print(f"설정 파일을 만들었습니다: {DEFAULT_CONFIG}")
        print("메일 수신자와 추가 사이트는 config.json에서 수정하면 됩니다.")


def read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def public_item(item: Dict[str, Any]) -> Dict[str, Any]:
    excluded = {"sort_dt", "event_start_ord"}
    return {key: value for key, value in item.items() if key not in excluded}


def append_history(path: Path, run_id: str, items: List[Dict[str, Any]]) -> None:
    if not items:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        for item in items:
            record = public_item(item)
            record["run_id"] = run_id
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def make_item_id(title: str, link: str) -> str:
    return sha256_text(f"{clean_text(title)}|{clean_text(link)}")


def fetch_url(url: str, timeout: int = 25) -> requests.Response:
    response = SESSION.get(url, timeout=timeout)
    response.raise_for_status()
    return response


def fetch_url_optional(url: str, timeout: int = 25) -> Optional[requests.Response]:
    try:
        return fetch_url(url, timeout=timeout)
    except Exception:
        return None


def first_text(data: Dict[str, Any], keys: Iterable[str]) -> str:
    for key in keys:
        value = clean_text(data.get(key, ""))
        if value:
            return value
    return ""


def parse_any_date(value: str) -> Optional[date]:
    value = clean_text(value).replace("/", "-").replace(".", "-")
    match = re.search(r"(20\d{2})-(\d{1,2})-(\d{1,2})", value)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    match = re.search(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일", value)
    if match:
        year = datetime.now(KST).year
        try:
            return date(year, int(match.group(1)), int(match.group(2)))
        except ValueError:
            return None
    return None


def parse_date_range(*texts: str) -> Tuple[Optional[date], Optional[date]]:
    merged = " ".join(clean_text(t) for t in texts if clean_text(t))
    normalized = merged.replace("/", "-").replace(".", "-")
    matches = re.findall(r"20\d{2}-\d{1,2}-\d{1,2}", normalized)
    parsed = [parse_any_date(m) for m in matches]
    parsed = [d for d in parsed if d is not None]
    if len(parsed) >= 2:
        return parsed[0], parsed[1]
    if len(parsed) == 1:
        return parsed[0], parsed[0]

    from_to_match = re.search(
        r"(\d{1,2})\s*월\s*(\d{1,2})\s*일\s*(?:부터|~|-|∼)\s*"
        r"(?:(\d{1,2})\s*월\s*)?(\d{1,2})\s*일?\s*(?:까지)?",
        merged,
    )
    if from_to_match:
        year = datetime.now(KST).year
        start_month = int(from_to_match.group(1))
        start_day = int(from_to_match.group(2))
        end_month = int(from_to_match.group(3) or start_month)
        end_day = int(from_to_match.group(4))
        try:
            return date(year, start_month, start_day), date(year, end_month, end_day)
        except ValueError:
            pass

    range_match = re.search(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일\s*[~\-∼]\s*(\d{1,2})\s*일", merged)
    if range_match:
        year = datetime.now(KST).year
        month = int(range_match.group(1))
        try:
            return date(year, month, int(range_match.group(2))), date(year, month, int(range_match.group(3)))
        except ValueError:
            pass

    single = parse_any_date(merged)
    if single:
        return single, single
    return None, None


def parse_datetime_value(value: str) -> Optional[datetime]:
    value = clean_text(value)
    if not value:
        return None
    try:
        parsed = parsedate_to_datetime(value)
        if parsed:
            return parsed.astimezone(KST).replace(tzinfo=None)
    except Exception:
        pass
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def short_date(value: Optional[date]) -> str:
    if value is None:
        return ""
    return f"{value.month}.{value.day}({WEEKDAYS[value.weekday()]})"


def period_text(start: Optional[date], end: Optional[date], fallback: str) -> str:
    if start and end:
        left = f"{start:%Y.%m.%d}.({WEEKDAYS[start.weekday()]})"
        right = f"{end:%Y.%m.%d}.({WEEKDAYS[end.weekday()]})"
        if start == end:
            return left
        return f"{left}~{right}"
    return clean_text(fallback) or "(자료 없음)"


def split_sentences(text: str) -> List[str]:
    text = clean_text(text)
    if not text:
        return []
    return [clean_text(p) for p in re.split(r"(?<=[.!?。])\s+|\n+", text) if clean_text(p)]


def summarize(text: str, max_len: int = 180) -> str:
    sentences = split_sentences(text)
    summary = " ".join(sentences[:2]) if sentences else clean_text(text)
    if not summary:
        return "(자료 없음)"
    if len(summary) > max_len:
        return summary[: max_len - 1].rstrip() + "…"
    return summary


def extract_place(text: str) -> str:
    text = clean_text(text)
    venue_suffixes = (
        "시청|구청|군청|문화회관|체육관|센터|공원|광장|호텔|역|마당|대학|"
        "컨벤션|아트홀|전시장|복합문화공간|수목원|체험관|운동장|경기장|"
        "해수욕장|해변|항|부두|시장|거리|로터리|스타디움|박물관|미술관|"
        "예술회관|문화센터|체육센터|마을|들판|일원|광장|야외무대|특설무대"
    )
    patterns = [
        r"(?:장소|위치|개최지|집결지|행사장)\s*[:：]?\s*([^\n\r,;|]+)",
        rf"까지\s*([가-힣A-Za-z0-9·ㆍ\-\s]+?(?:{venue_suffixes}))\s*에서",
        rf"([가-힣A-Za-z0-9·ㆍ\-\s]+?(?:{venue_suffixes}))\s*에서\s*[\"'‘“]?[가-힣A-Za-z0-9·ㆍ\-\s]*(?:축제|행사|집회|시위|공연|전시|박람회|콘서트|마라톤|대회)",
        rf"([가-힣A-Za-z0-9·ㆍ\-\s]+?(?:{venue_suffixes}))",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            place = clean_text(match.group(1)).strip(" ,.|:：")
            place = re.sub(r"^\d{1,2}\s*일?\s*까지\s*", "", place).strip()
            place = re.sub(r"^(?:오는|올해|내달|다음달|이번|오는\s+\d{1,2}\s*월)\s+", "", place).strip()
            if len(place) >= 2 and not is_weak_place_name(place):
                return place
    return ""

GEOCODER_PLACE_CACHE: Dict[str, str] = {}
AI_EXTRACTION_CACHE: Dict[str, Dict[str, Any]] = {}
AI_EXTRACTION_WARNING_PRINTED = False
ARTICLE_DETAIL_STATS: Dict[str, int] = {
    "attempted": 0,
    "success": 0,
    "failed": 0,
    "empty": 0,
}
AI_EXTRACTION_STATS: Dict[str, int] = {
    "attempted": 0,
    "success": 0,
    "failed": 0,
    "skipped_disabled": 0,
    "skipped_not_rss": 0,
    "skipped_short_body": 0,
    "skipped_no_key": 0,
    "fallback_for_missing": 0,
    "non_event": 0,
}
GEOCODER_STATS: Dict[str, int] = {
    "attempted": 0,
    "success": 0,
    "failed": 0,
    "skipped_no_key": 0,
    "skipped_address": 0,
    "skipped_weak": 0,
    "skipped_disabled": 0,
}


def looks_like_address(value: str) -> bool:
    value = clean_text(value)
    if not value or value == "(자료 없음)":
        return False

    patterns = [
        r"(?:로|길|대로|번길)\s*\d+(?:-\d+)?",
        r"(?:읍|면|동|리)\s*\d+(?:-\d+)?",
        r"(?:시|군|구)\s+.+(?:로|길|대로|번길)\s*\d+(?:-\d+)?",
    ]

    return any(re.search(pattern, value) for pattern in patterns)


def looks_like_admin_area_only(value: str) -> bool:
    value = clean_text(value).strip(" ,.|:：")
    if not value:
        return False
    if looks_like_address(value):
        return False
    if re.fullmatch(r"(?:부산|울산|경남|경상남도|부산광역시|울산광역시)(?:시|도)?", value):
        return True
    if re.fullmatch(r"(?:부산|울산|경남|경상남도|부산광역시|울산광역시)?\s*[가-힣]+(?:시|군|구|읍|면|동)", value):
        return True
    if re.fullmatch(r"(?:부산|울산|경남|경상남도|부산광역시|울산광역시)?\s*[가-힣]+(?:시|군|구)\s+[가-힣]+(?:읍|면|동)?", value):
        return True
    return False


def is_weak_place_name(value: str) -> bool:
    value = clean_text(value)

    if not value:
        return True

    weak_values = {
        "(자료 없음)",
        "지역",
        "행사",
        "축제",
        "문화",
        "공연",
        "부산",
        "울산",
        "경남",
        "경상남도",
    }

    if value in weak_values:
        return True

    if looks_like_admin_area_only(value):
        return True

    if len(value) <= 2:
        return True

    # 너무 긴 문장 조각은 장소명 검색 대상에서 제외
    if len(value) > 40 and not looks_like_address(value):
        return True

    bad_fragments = [
        "지역경제",
        "지역사회",
        "다문화 가정",
        "복지위기가구",
        "기업사랑 시민축제로 지역",
        "손잡고 지역",
        "찾아가는",
        "발굴",
        "협약",
        "도움",
        "나눔",
        "대상",
        "개최 알림",
        "우수사례",
        "후보",
        "시장 후보",
        "도지사",
        "개혁신당",
        "국민의힘",
        "더불어민주당",
        "민주당",
        "선거사무소",
    ]

    if any(fragment in value for fragment in bad_fragments):
        return True

    return False


def resolve_place_to_address_tmap(place_name: str, region_hint: str = "") -> str:
    place_name = clean_text(place_name)

    if is_weak_place_name(place_name):
        GEOCODER_STATS["skipped_weak"] += 1
        return ""

    if looks_like_address(place_name):
        GEOCODER_STATS["skipped_address"] += 1
        return place_name

    app_key = os.getenv("TMAP_APP_KEY", "").strip()
    if not app_key:
        GEOCODER_STATS["skipped_no_key"] += 1
        print("TMAP_APP_KEY 없음 -> 장소 주소 변환 생략")
        return ""

    GEOCODER_STATS["attempted"] += 1
    query_candidates = []

    if region_hint:
        for region in region_hint.split():
            query_candidates.append(clean_text(f"{region} {place_name}"))

    query_candidates.append(place_name)

    url = "https://apis.openapi.sk.com/tmap/pois"

    headers = {
        "appKey": app_key,
        "Accept": "application/json",
    }

    for query in query_candidates:
        if not query:
            continue

        if query in GEOCODER_PLACE_CACHE:
            cached = GEOCODER_PLACE_CACHE[query]
            if cached:
                return cached
            continue

        params = {
            "version": "1",
            "format": "json",
            "searchKeyword": query,
            "count": "3",
            "searchType": "all",
            "resCoordType": "WGS84GEO",
            "reqCoordType": "WGS84GEO",
        }

        try:
            response = SESSION.get(url, headers=headers, params=params, timeout=10)

            if response.status_code == 204:
                print(f"Tmap 장소 검색 결과 없음: {query}")
                GEOCODER_PLACE_CACHE[query] = ""
                continue

            if response.status_code != 200:
                print(f"Tmap API 상태코드: {response.status_code}")
                print(f"Tmap API 응답: {response.text[:500]}")

            response.raise_for_status()

            try:
                data = response.json()
            except Exception as e:
                print(f"Tmap JSON 파싱 실패: {query} / {e}")
                GEOCODER_PLACE_CACHE[query] = ""
                continue

            pois = data.get("searchPoiInfo", {}).get("pois", {}).get("poi", [])
            if not pois:
                print(f"Tmap 장소 검색 결과 없음: {query}")
                GEOCODER_PLACE_CACHE[query] = ""
                continue

            for poi in pois:
                name = clean_text(poi.get("name", ""))

                upper_addr = clean_text(poi.get("upperAddrName", ""))
                middle_addr = clean_text(poi.get("middleAddrName", ""))
                lower_addr = clean_text(poi.get("lowerAddrName", ""))
                detail_addr = clean_text(poi.get("detailAddrName", ""))

                road_name = clean_text(poi.get("roadName", ""))
                first_build_no = clean_text(poi.get("firstBuildNo", ""))
                second_build_no = clean_text(poi.get("secondBuildNo", ""))

                build_no = first_build_no
                if second_build_no and second_build_no != "0":
                    build_no = f"{first_build_no}-{second_build_no}"

                road_address = clean_text(
                    " ".join(
                        part for part in [
                            upper_addr,
                            middle_addr,
                            road_name,
                            build_no,
                        ]
                        if part
                    )
                )

                jibun_address = clean_text(
                    " ".join(
                        part for part in [
                            upper_addr,
                            middle_addr,
                            lower_addr,
                            detail_addr,
                        ]
                        if part
                    )
                )

                address = road_address or jibun_address

                if address:
                    if name and name not in address:
                        address = f"{address} ({name})"

                    GEOCODER_PLACE_CACHE[query] = address
                    GEOCODER_STATS["success"] += 1
                    return address

            GEOCODER_PLACE_CACHE[query] = ""

        except Exception as e:
            print(f"Tmap 장소 주소 변환 실패: {query} / {e}")
            GEOCODER_PLACE_CACHE[query] = ""
            continue

    GEOCODER_STATS["failed"] += 1
    return ""


def enrich_place_with_geocoder(place: str, config: Dict[str, Any]) -> str:
    place = clean_text(place)

    if not place or place == "(자료 없음)":
        return place

    if looks_like_address(place):
        return place

    geocoding = config.get("geocoding", {})
    if not geocoding.get("enabled", False):
        GEOCODER_STATS["skipped_disabled"] += 1
        return place

    provider = clean_text(geocoding.get("provider", "tmap")).lower()
    region_hint = clean_text(geocoding.get("region_hint", "부산 울산 경남"))

    resolved = ""

    if provider == "tmap":
        resolved = resolve_place_to_address_tmap(place, region_hint=region_hint)
    else:
        print(f"지원하지 않는 geocoding provider: {provider}")

    if resolved:
        print(f"장소 주소 변환: {place} -> {resolved}")
        return resolved

    return place



def extract_crowd(text: str) -> Tuple[Optional[int], str]:
    text = clean_text(text)
    patterns = [
        r"(\d{1,3}(?:,\d{3})*|\d+)\s*(?:명|여명)",
        r"(\d+(?:\.\d+)?)\s*만\s*(?:명|여명)",
        r"(\d+(?:\.\d+)?)\s*천\s*(?:명|여명)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        raw = match.group(1)
        if "만" in match.group(0):
            value = int(float(raw.replace(",", "")) * 10000)
        elif "천" in match.group(0):
            value = int(float(raw.replace(",", "")) * 1000)
        else:
            value = int(raw.replace(",", ""))
        return value, f"{value:,}"
    return None, "(자료 없음)"


def extraction_label(source: str, value: str = "") -> str:
    source = clean_text(source)
    value = clean_text(value)
    tmap_suffix = "+Tmap" if source.endswith("_tmap") else ""
    source = source.replace("_tmap", "")
    if source == "ai":
        return f"AI{tmap_suffix}"
    if source == "rule":
        return f"Rule 약함{tmap_suffix}" if value and is_weak_place_name(value) else f"Rule{tmap_suffix}"
    return "없음"


def build_extraction_note(
    ai_result: Dict[str, Any],
    period_source: str,
    place_source: str,
    crowd_source: str,
    place_value: str,
) -> str:
    ai_used = any(clean_text(source).startswith("ai") for source in [period_source, place_source, crowd_source])
    if ai_result and ai_used:
        overall = "AI보완"
    elif ai_result:
        overall = "AI확인(보완없음)"
    else:
        overall = "Rule-Based"
    confidence = ""
    if ai_result and ai_result.get("confidence") not in {None, ""}:
        try:
            confidence = f" / 신뢰도:{float(ai_result.get('confidence')):.2f}"
        except (TypeError, ValueError):
            confidence = ""
    parts = [
        f"전체:{overall}{confidence}",
        f"기간:{extraction_label(period_source)}",
        f"장소:{extraction_label(place_source, place_value)}",
        f"인원:{extraction_label(crowd_source)}",
    ]
    return " / ".join(parts)


def parse_iso_date(value: str) -> Optional[date]:
    value = clean_text(value)
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def find_json_text(data: Any) -> str:
    if isinstance(data, dict):
        if isinstance(data.get("text"), str):
            return data["text"]
        for value in data.values():
            found = find_json_text(value)
            if found:
                return found
    if isinstance(data, list):
        for value in data:
            found = find_json_text(value)
            if found:
                return found
    return ""


def extract_event_with_ai(raw: Dict[str, Any], config: Dict[str, Any], text: str, force: bool = False) -> Dict[str, Any]:
    global AI_EXTRACTION_WARNING_PRINTED

    ai_cfg = config.get("ai_extraction", {})
    if not ai_cfg.get("enabled", False):
        AI_EXTRACTION_STATS["skipped_disabled"] += 1
        return {}

    if not force and ai_cfg.get("only_for_rss", True) and clean_text(raw.get("source_type")).lower() != "rss":
        AI_EXTRACTION_STATS["skipped_not_rss"] += 1
        return {}

    if not force and len(clean_text(raw.get("body"))) < int(ai_cfg.get("min_body_chars", 200)):
        AI_EXTRACTION_STATS["skipped_short_body"] += 1
        return {}

    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key:
        AI_EXTRACTION_STATS["skipped_no_key"] += 1
        if not AI_EXTRACTION_WARNING_PRINTED:
            print("OPENAI_API_KEY 없음 -> AI 본문 추출 생략")
            AI_EXTRACTION_WARNING_PRINTED = True
        return {}

    cache_key = sha256_text(f"{raw.get('title', '')}|{raw.get('link', '')}|{text[:4000]}|force:{force}")
    if cache_key in AI_EXTRACTION_CACHE:
        return AI_EXTRACTION_CACHE[cache_key]

    AI_EXTRACTION_STATS["attempted"] += 1
    today = datetime.now(KST).date().isoformat()
    model = clean_text(ai_cfg.get("model", "gpt-4o-mini")) or "gpt-4o-mini"
    prompt = f"""
아래 기사를 읽고 사람이 실제로 모일 수 있는 행사/축제/공연/전시/집회/시위 정보를 JSON으로만 추출해 주세요.

판단 기준:
- 단순 봉사, 교육, 채용, 포상, 홍보, 기관 내부 기념식은 is_event=false에 가깝습니다.
- 장소는 기사에 나온 실제 개최 장소를 가장 구체적으로 적습니다.
- 날짜가 월/일만 있으면 오늘 날짜({today}) 기준의 연도로 보정합니다.
- 모르면 빈 문자열 또는 null을 넣습니다.
- expected_crowd는 기사에 명시된 인원만 숫자로 넣고, 추정하지 않습니다.

제목: {clean_text(raw.get("title"))}
게시일: {clean_text(raw.get("published"))}
본문:
{clean_text(text)[:6000]}
""".strip()

    schema = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "is_event": {"type": "boolean"},
            "event_name": {"type": "string"},
            "start_date": {"type": "string"},
            "end_date": {"type": "string"},
            "place": {"type": "string"},
            "event_type": {
                "type": "string",
                "enum": ["행사", "축제", "공연", "전시", "스포츠", "체험", "집회", "기타"],
            },
            "expected_crowd": {"type": ["integer", "null"]},
            "confidence": {"type": "number"},
            "reason": {"type": "string"},
        },
        "required": [
            "is_event",
            "event_name",
            "start_date",
            "end_date",
            "place",
            "event_type",
            "expected_crowd",
            "confidence",
            "reason",
        ],
    }

    payload = {
        "model": model,
        "input": [
            {
                "role": "system",
                "content": "당신은 이동통신 특별소통 대응을 위한 행사 정보 추출 도우미입니다. 반드시 JSON 스키마에 맞춰 답합니다.",
            },
            {"role": "user", "content": prompt},
        ],
        "text": {
            "format": {
                "type": "json_schema",
                "name": "event_extraction",
                "strict": True,
                "schema": schema,
            }
        },
    }

    try:
        response = SESSION.post(
            "https://api.openai.com/v1/responses",
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=30,
        )
        response.raise_for_status()
        data = response.json()
        result_text = find_json_text(data)
        result = json.loads(result_text) if result_text else {}
        if isinstance(result, dict):
            AI_EXTRACTION_CACHE[cache_key] = result
            AI_EXTRACTION_STATS["success"] += 1
            if result.get("is_event") is False:
                AI_EXTRACTION_STATS["non_event"] += 1
            return result
    except Exception as exc:
        AI_EXTRACTION_STATS["failed"] += 1
        if not AI_EXTRACTION_WARNING_PRINTED:
            print(f"AI 본문 추출 실패 -> 룰 기반 추출로 진행: {exc}")
            AI_EXTRACTION_WARNING_PRINTED = True

    return {}


def classify_type(text: str) -> str:
    text = clean_text(text)
    rules = [
        ("집회", ["집회", "시위", "기자회견", "행진", "규탄", "총궐기"]),
        ("축제", ["축제", "페스티벌", "불꽃", "개막"]),
        ("공연", ["공연", "콘서트", "버스킹", "초대가수"]),
        ("전시", ["전시", "박람회", "엑스포"]),
        ("스포츠", ["걷기", "마라톤", "대회", "리그", "투어", "에어쇼", "체육대회"]),
        ("체험", ["체험", "캠프", "교육"]),
    ]
    for label, keywords in rules:
        if any(keyword in text for keyword in keywords):
            return label
    return "행사"


def grade_event(crowd: Optional[int], rules: Dict[str, Any]) -> str:
    if crowd is None:
        return clean_text(rules.get("unknown_crowd_grade", "검토")) or "검토"
    thresholds = rules.get("grade_thresholds", {})
    high = int(thresholds.get("상", 10000))
    mid = int(thresholds.get("중", 3000))
    if crowd >= high:
        return "상"
    if crowd >= mid:
        return "중"
    return "하"


def priority_score(
    start: Optional[date],
    place: str,
    crowd: Optional[int],
    event_type: str,
    source_type: str,
    grade: str,
) -> int:
    score = 0
    if start:
        score += 45
    if place and place != "(자료 없음)":
        score += 20
    if crowd is not None:
        score += 25
    if source_type in {"gyeongnam_festa", "busan_notice", "html"}:
        score += 20
    if event_type in {"축제", "공연", "집회", "스포츠", "전시"}:
        score += 15
    if grade == "검토":
        score -= 5
    return score


def classify_feature(start: Optional[date], end: Optional[date], text: str) -> str:
    if start and end and end > start:
        return "1일이상"
    if re.search(r"\d{1,2}\s*[:시]\s*\d{0,2}\s*[~\-∼]\s*\d{1,2}\s*[:시]", text):
        return "1시간이상"
    if re.search(r"\(\s*\d{1,2}\s*:\s*\d{2}\s*[~\-∼]\s*\d{1,2}\s*:\s*\d{2}\s*\)", text):
        return "1시간이상"
    return "(자료 없음)"


def contains_any(text: str, keywords: Iterable[str]) -> bool:
    return any(keyword in text for keyword in keywords if keyword)


def keyword_hit_count(text: str, keywords: Iterable[str]) -> int:
    return sum(1 for keyword in keywords if keyword and keyword in text)


def normalize_rss_title(title: str) -> str:
    title = clean_text(title)
    if not title:
        return title

    # RSS 제목 뒤에 붙는 매체명/도메인 꼬리를 최대한 제거해 중복 노이즈를 줄인다.
    parts = re.split(r"\s[-–—]\s", title)
    if len(parts) < 2:
        return title

    left = clean_text(parts[0])
    right = clean_text(parts[-1]).lower()
    if not left or not right:
        return title

    source_suffix_markers = (
        ".com", ".co.kr", ".net", ".kr", "뉴스", "일보", "신문", "경제",
        "매거진", "미디어", "브리핑", "연합뉴스", "daum", "nate", "v.daum.net",
    )
    if len(right) <= 40 and (
        any(marker in right for marker in source_suffix_markers) or re.fullmatch(r"[a-z0-9.&()/_-]+", right)
    ):
        return left

    return title


def extract_article_text(html: str, max_chars: int = 6000) -> str:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup.select("script, style, noscript, header, footer, nav, aside, form"):
        tag.decompose()

    chunks: List[str] = []
    for selector in (
        'meta[property="og:description"]',
        'meta[name="description"]',
    ):
        meta = soup.select_one(selector)
        content = clean_text(meta.get("content", "") if meta else "")
        if content:
            chunks.append(content)

    candidates = soup.select(
        "article, [itemprop='articleBody'], #articleBody, #news_body_area, "
        ".article_body, .article-body, .article-view, .news_view, .view_con, .content"
    )
    if not candidates:
        candidates = soup.find_all("p")

    for candidate in candidates[:30]:
        text = clean_text(candidate.get_text(" ", strip=True))
        if len(text) >= 30:
            chunks.append(text)

    merged = clean_text(" ".join(chunks))
    if len(merged) > max_chars:
        return merged[:max_chars].rstrip()
    return merged


def fetch_article_detail(link: str, rss_cfg: Dict[str, Any]) -> Tuple[str, str]:
    link = clean_text(link)
    if not link:
        return "", ""

    ARTICLE_DETAIL_STATS["attempted"] += 1
    timeout = int(rss_cfg.get("article_timeout_seconds", 8))
    max_chars = int(rss_cfg.get("article_max_chars", 6000))
    response = fetch_url_optional(link, timeout=timeout)
    if response is None:
        ARTICLE_DETAIL_STATS["failed"] += 1
        return "", ""

    content_type = response.headers.get("content-type", "").lower()
    if "html" not in content_type and "<html" not in response.text[:500].lower():
        ARTICLE_DETAIL_STATS["empty"] += 1
        return "", clean_text(response.url)

    article_text = extract_article_text(response.text, max_chars=max_chars)
    if article_text:
        ARTICLE_DETAIL_STATS["success"] += 1
    else:
        ARTICLE_DETAIL_STATS["empty"] += 1
    return article_text, clean_text(response.url)


def rss_signal_score(item: Dict[str, Any], filters: Dict[str, Any]) -> int:
    rss_cfg = filters.get("rss", {})
    title = clean_text(item.get("title", ""))
    body = clean_text(item.get("body", ""))
    text = f"{title} {body}"

    strong_keywords = [clean_text(v) for v in rss_cfg.get("strong_keywords", []) if clean_text(v)]
    weak_keywords = [clean_text(v) for v in rss_cfg.get("weak_keywords", []) if clean_text(v)]
    venue_keywords = [clean_text(v) for v in rss_cfg.get("venue_keywords", []) if clean_text(v)]
    protest_keywords = [clean_text(v) for v in rss_cfg.get("protest_keywords", []) if clean_text(v)]

    score = 0

    # 제목에 들어간 신호를 조금 더 강하게 반영한다.
    has_title_strong = contains_any(title, strong_keywords)
    has_body_strong = contains_any(body, strong_keywords)
    has_protest_signal = contains_any(text, protest_keywords)

    if has_title_strong:
        score += 3
    if has_body_strong:
        score += 2
    if has_protest_signal:
        score += 4
    if contains_any(text, venue_keywords):
        score += 1
    if re.search(r"\d{1,3}(?:,\d{3})*\s*(?:명|여명)", text):
        score += 2
    if re.search(r"\d+(?:\.\d+)?\s*(?:만|천)\s*(?:명|여명)", text):
        score += 2
    if re.search(r"(20\d{2}|19\d{2})[.\-/]\d{1,2}[.\-/]\d{1,2}", text) or re.search(r"\d{1,2}\s*월\s*\d{1,2}\s*일", text):
        score += 1
    if contains_any(text, weak_keywords) and not (has_title_strong or has_body_strong or has_protest_signal):
        score -= min(6, keyword_hit_count(text, weak_keywords))

    return score


def passes_filters(item: Dict[str, Any], filters: Dict[str, Any]) -> bool:
    source_type = clean_text(item.get("source_type")).lower()
    item_text = f"{item.get('title', '')} {item.get('body', '')} {item.get('place', '')}"
    text = f"{item.get('source_name', '')} {item_text}"
    include = [clean_text(v) for v in filters.get("include_keywords", []) if clean_text(v)]
    exclude = [clean_text(v) for v in filters.get("exclude_keywords", []) if clean_text(v)]
    regions = [clean_text(v) for v in filters.get("region_keywords", []) if clean_text(v)]
    if any(keyword in text for keyword in exclude):
        return False

    if source_type == "rss":
        rss_cfg = filters.get("rss", {})
        min_score = int(rss_cfg.get("min_signal_score", 4))
        score = rss_signal_score(item, filters)
        if score < min_score:
            return False
        if filters.get("require_region", False) and regions and not any(keyword in item_text for keyword in regions):
            return False
        return True

    if filters.get("require_region", False) and regions:
        if not any(keyword in text for keyword in regions):
            return False
    if filters.get("require_keyword", True) and include:
        return any(keyword in text for keyword in include)
    return True


def is_expired(start: Optional[date], end: Optional[date], filters: Dict[str, Any]) -> bool:
    if not filters.get("skip_expired_events", True):
        return False
    if end is None:
        return False
    return end < datetime.now(KST).date()


def parse_rss_source(source: Dict[str, Any], rss_cfg: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    rss_cfg = rss_cfg or {}
    response = fetch_url(source["url"])
    root = ET.fromstring(response.content)
    rows: List[Dict[str, Any]] = []
    fetch_detail = bool(source.get("fetch_article_detail", rss_cfg.get("fetch_article_detail", True)))
    max_fetch = int(source.get("max_article_fetch_per_source", rss_cfg.get("max_article_fetch_per_source", 15)))
    fetched_count = 0

    for element in root.iter():
        local = element.tag.rsplit("}", 1)[-1].lower()
        if local not in {"item", "entry"}:
            continue

        fields: Dict[str, str] = {}
        for child in element:
            key = child.tag.rsplit("}", 1)[-1].lower()
            if child.text:
                fields[key] = clean_text(child.text)
            if key == "link" and not fields.get("link"):
                fields["link"] = clean_text(child.attrib.get("href", ""))

        title = normalize_rss_title(fields.get("title", ""))
        body = fields.get("description") or fields.get("summary") or fields.get("content") or ""
        link = fields.get("link", source["url"])
        detail_text = ""
        resolved_link = ""
        if fetch_detail and fetched_count < max_fetch:
            detail_text, resolved_link = fetch_article_detail(link, rss_cfg)
            fetched_count += 1
        if detail_text and detail_text not in body:
            body = clean_text(f"{body} {detail_text}")
        rows.append(
            {
                "source_name": source["name"],
                "source_type": "rss",
                "source_url": source["url"],
                "title": title,
                "link": resolved_link or link,
                "published": fields.get("pubdate") or fields.get("published") or fields.get("updated") or "",
                "event_period": "",
                "body": body,
                "place": "",
            }
        )
    return rows


def parse_html_source(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    response = fetch_url(source["url"])
    soup = BeautifulSoup(response.text, "html.parser")
    containers = soup.select(source.get("item_selector", "article, li, .item, .news-item, .post, .card"))
    if not containers:
        containers = [soup]
    rows: List[Dict[str, Any]] = []
    for container in containers:
        title_el = container.select_one(source.get("title_selector", "a, h2, h3, h4"))
        link_el = container.select_one(source.get("link_selector", "a"))
        body_el = container.select_one(source.get("body_selector", "p, .desc, .summary, .content"))
        date_el = container.select_one(source.get("date_selector", "time, .date, .day"))
        title = clean_text(title_el.get_text(" ", strip=True) if title_el else "")
        body = clean_text(body_el.get_text(" ", strip=True) if body_el else "")
        link = ""
        if link_el:
            link = clean_text(link_el.get("href") or link_el.get("data-href") or "")
        if not title and not body:
            continue
        rows.append(
            {
                "source_name": source["name"],
                "source_type": "html",
                "source_url": source["url"],
                "title": title or "(제목 없음)",
                "link": urljoin(source["url"], link) if link else source["url"],
                "published": clean_text(date_el.get_text(" ", strip=True) if date_el else ""),
                "event_period": "",
                "body": body,
                "place": "",
            }
        )
    return rows


def parse_busan_notice(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    response = fetch_url(source["url"])
    soup = BeautifulSoup(response.text, "html.parser")
    rows: List[Dict[str, Any]] = []
    for tr in soup.select("table tbody tr"):
        cells = tr.find_all("td")
        link_tag = tr.find("a", href=True)
        if not link_tag or len(cells) < 2:
            continue
        title = clean_text(link_tag.get_text(" ", strip=True))
        texts = [clean_text(td.get_text(" ", strip=True)) for td in cells if clean_text(td.get_text(" ", strip=True))]
        published = ""
        department = ""
        for idx, text in enumerate(texts):
            if re.search(r"20\d{2}[-.]\d{1,2}[-.]\d{1,2}", text):
                published = text
                if idx > 0:
                    department = texts[idx - 1]
                break
        rows.append(
            {
                "source_name": source["name"],
                "source_type": "busan_notice",
                "source_url": source["url"],
                "title": title,
                "link": urljoin(source["url"], link_tag["href"]),
                "published": published,
                "event_period": "",
                "body": department,
                "place": "",
            }
        )
    return rows


def find_first_list(data: Any) -> List[Dict[str, Any]]:
    if isinstance(data, list):
        return [x for x in data if isinstance(x, dict)]
    if isinstance(data, dict):
        for key in ("resultData", "data", "list", "items", "result", "body"):
            if key in data:
                found = find_first_list(data[key])
                if found:
                    return found
        for value in data.values():
            found = find_first_list(value)
            if found:
                return found
    if isinstance(data, str):
        try:
            return find_first_list(json.loads(data))
        except Exception:
            return []
    return []


def parse_gyeongnam_festa(source: Dict[str, Any]) -> List[Dict[str, Any]]:
    api_url = urljoin(source["url"], "/api/callFestivalList.do")
    response = fetch_url(api_url)
    festivals = find_first_list(response.json())
    rows: List[Dict[str, Any]] = []
    for festival in festivals:
        title = first_text(
            festival,
            ["siteName", "festivalName", "eventName", "title", "name", "contentName"],
        )
        if not title:
            continue
        start = first_text(festival, ["festivalStartDate", "startDate", "eventStartDate", "beginDate"])
        end = first_text(festival, ["festivalEndDate", "endDate", "eventEndDate", "finishDate"])
        place = first_text(
            festival,
            [
                "festivalAddress", "festivalDetailAddress", "address", "addr",
                "roadAddr", "jibunAddr", "place", "placeNm", "eventPlace",
                "festivalPlace", "location", "venue", "sigunguName"
            ],
        )
        detail_place = first_text(festival, ["festivalDetailAddress"])
        if place and detail_place and detail_place not in place:
            place = f"{place} {detail_place}"
        body = first_text(
            festival,
            ["contentsIntro", "siteKeyword", "content", "summary", "description", "intro", "mainContent"],
        )
        link = first_text(festival, ["linkUrl", "url", "homepage", "homepageUrl"])
        if not link:
            sub_path = first_text(festival, ["subPath", "path"])
            link = urljoin(source["url"], "/" + sub_path.lstrip("/")) if sub_path else source["url"]
        rows.append(
            {
                "source_name": source["name"],
                "source_type": "gyeongnam_festa",
                "source_url": source["url"],
                "title": title,
                "link": link,
                "published": "",
                "event_period": f"{start} ~ {end}".strip(" ~"),
                "body": body,
                "place": place,
            }
        )
    return rows


def parse_source(source: Dict[str, Any], config: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    source_type = clean_text(source.get("type", "rss")).lower()
    if source_type == "rss":
        rss_cfg = ((config or {}).get("filters", {}) or {}).get("rss", {})
        return parse_rss_source(source, rss_cfg)
    if source_type == "html":
        return parse_html_source(source)
    if source_type == "busan_notice":
        return parse_busan_notice(source)
    if source_type == "gyeongnam_festa":
        return parse_gyeongnam_festa(source)
    raise ValueError(f"지원하지 않는 source type: {source_type}")


def normalize_item(raw: Dict[str, Any], config: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    filters = config.get("filters", {})
    rules = config.get("rules", {})
    if not passes_filters(raw, filters):
        return None

    text = " ".join(
        [
            clean_text(raw.get("title")),
            clean_text(raw.get("body")),
            clean_text(raw.get("place")),
            clean_text(raw.get("event_period")),
        ]
    )
    ai_result = extract_event_with_ai(raw, config, text)
    if (
        ai_result
        and ai_result.get("is_event") is False
        and config.get("ai_extraction", {}).get("drop_non_events", False)
        and float(ai_result.get("confidence") or 0) >= 0.75
    ):
        return None

    start, end = parse_date_range(raw.get("event_period", ""), text)
    period_source = "rule" if start or end else "none"
    rule_place = clean_text(raw.get("place")) or extract_place(text)
    needs_ai_fallback = (
        not ai_result
        and config.get("ai_extraction", {}).get("fallback_on_missing", True)
        and (
            start is None
            or end is None
            or not rule_place
            or is_weak_place_name(rule_place)
        )
    )
    if needs_ai_fallback:
        AI_EXTRACTION_STATS["fallback_for_missing"] += 1
        ai_result = extract_event_with_ai(raw, config, text, force=True)

    ai_start = parse_iso_date(clean_text(ai_result.get("start_date", ""))) if ai_result else None
    ai_end = parse_iso_date(clean_text(ai_result.get("end_date", ""))) if ai_result else None
    if ai_start and (start is None or float(ai_result.get("confidence") or 0) >= 0.6):
        start = ai_start
        period_source = "ai"
    if ai_end and (end is None or float(ai_result.get("confidence") or 0) >= 0.6):
        end = ai_end
        period_source = "ai"
    if start and end is None:
        end = start

    if is_expired(start, end, filters):
        return None

    ai_place = clean_text(ai_result.get("place", "")) if ai_result else ""
    raw_place = rule_place
    place_source = "rule" if raw_place else "none"
    if ai_place and not is_weak_place_name(ai_place) and (not raw_place or is_weak_place_name(raw_place)):
        raw_place = ai_place
        place_source = "ai"
    raw_place = raw_place or "(자료 없음)"
    place = enrich_place_with_geocoder(raw_place, config)
    if place != raw_place and place_source in {"rule", "ai"}:
        place_source = f"{place_source}_tmap"
    crowd_value, crowd_display = extract_crowd(text)
    crowd_source = "rule" if crowd_value is not None else "none"
    ai_crowd = ai_result.get("expected_crowd") if ai_result else None
    if crowd_value is None and isinstance(ai_crowd, int):
        crowd_value = ai_crowd
        crowd_display = f"{ai_crowd:,}"
        crowd_source = "ai"
    event_type = classify_type(text)
    ai_type = clean_text(ai_result.get("event_type", "")) if ai_result else ""
    if ai_type in {"행사", "축제", "공연", "전시", "스포츠", "체험", "집회"}:
        if event_type == "행사" or float(ai_result.get("confidence") or 0) >= 0.6:
            event_type = ai_type
    grade = grade_event(crowd_value, rules)
    source_type = clean_text(raw.get("source_type"))
    score = priority_score(start, place, crowd_value, event_type, source_type, grade)
    generated_at = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")
    link = clean_text(raw.get("link")) or clean_text(raw.get("source_url"))
    original_title = clean_text(raw.get("title")) or "(제목 없음)"
    ai_title = clean_text(ai_result.get("event_name", "")) if ai_result else ""
    title = ai_title if ai_title and float(ai_result.get("confidence") or 0) >= 0.6 else original_title
    extraction_note = build_extraction_note(ai_result, period_source, place_source, crowd_source, raw_place)
    ai_used = any(clean_text(source).startswith("ai") for source in [period_source, place_source, crowd_source])
    status = "신규(AI보완)" if ai_result and ai_used else "신규(AI확인)" if ai_result else "신규"

    return {
        "id": make_item_id(title, link),
        "period": period_text(start, end, raw.get("event_period", "")),
        "title": title,
        "place": place,
        "type": event_type,
        "start_short": short_date(start),
        "end_short": short_date(end),
        "grade": grade,
        "feature": classify_feature(start, end, text),
        "network": clean_text(rules.get("default_network", "검토")) or "검토",
        "crowd": crowd_display,
        "crowd_value": crowd_value,
        "source_name": clean_text(raw.get("source_name")),
        "published": clean_text(raw.get("published")) or "(자료 없음)",
        "link": link,
        "summary": summarize(text),
        "collected_at": generated_at,
        "extraction_method": extraction_note,
        "status": status,
        "sort_dt": parse_datetime_value(raw.get("published", "")),
        "event_start_ord": start.toordinal() if start else 99999999,
        "priority_score": score,
    }


def sort_items(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def key(item: Dict[str, Any]) -> Tuple[int, int, int, float, str]:
        sort_dt = item.get("sort_dt")
        priority = int(item.get("priority_score", 0))
        event_start = int(item.get("event_start_ord", 99999999))
        if isinstance(sort_dt, datetime):
            return (-priority, 0 if event_start < 99999999 else 1, event_start, -sort_dt.timestamp(), item.get("title", ""))
        return (-priority, 0 if event_start < 99999999 else 1, event_start, 0.0, item.get("title", ""))

    return sorted(items, key=key)


def apply_limits(items: List[Dict[str, Any]], config: Dict[str, Any]) -> List[Dict[str, Any]]:
    output_cfg = config.get("output", {})
    total = int(output_cfg.get("max_items_total", 30))
    per_source = int(output_cfg.get("max_items_per_source", 10))
    limited: List[Dict[str, Any]] = []
    counts: Dict[str, int] = {}
    for item in sort_items(items):
        source_name = item.get("source_name", "")
        counts[source_name] = counts.get(source_name, 0)
        if per_source > 0 and counts[source_name] >= per_source:
            continue
        counts[source_name] += 1
        limited.append(item)
        if total > 0 and len(limited) >= total:
            break
    return limited


def dedupe_new_items(items: List[Dict[str, Any]], state: Dict[str, Any], include_seen: bool) -> List[Dict[str, Any]]:
    seen = state.setdefault("seen_ids", {})
    fresh: List[Dict[str, Any]] = []
    for item in items:
        item_id = item["id"]
        if include_seen or item_id not in seen:
            fresh.append(item)
        seen[item_id] = {
            "title": item["title"],
            "source": item["source_name"],
            "seen_at": datetime.now(KST).isoformat(timespec="seconds"),
        }
    if len(seen) > 1500:
        ordered = sorted(seen.items(), key=lambda kv: kv[1].get("seen_at", ""))
        state["seen_ids"] = dict(ordered[-1500:])
    state["last_run"] = datetime.now(KST).isoformat(timespec="seconds")
    return fresh


def collect_items(config: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], List[str]]:
    sources = config.get("sources", [])
    collected: List[Dict[str, Any]] = []
    errors: List[str] = []

    with ThreadPoolExecutor(max_workers=min(6, max(1, len(sources)))) as executor:
        futures = {executor.submit(parse_source, source, config): source for source in sources}
        for future in as_completed(futures):
            source = futures[future]
            try:
                for raw in future.result():
                    item = normalize_item(raw, config)
                    if item:
                        collected.append(item)
            except Exception as exc:
                errors.append(f"{source.get('name', '(이름 없음)')}: {exc}")

    unique: Dict[str, Dict[str, Any]] = {}
    for item in collected:
        unique[item["id"]] = item
    return list(unique.values()), errors


def col_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def cell_xml(row: int, col: int, value: Any, style: int = 0, numeric: bool = False) -> str:
    ref = f"{col_name(col)}{row}"
    style_attr = f' s="{style}"' if style else ""
    if value is None or value == "":
        return f'<c r="{ref}"{style_attr}/>'
    if numeric and isinstance(value, (int, float)):
        return f'<c r="{ref}"{style_attr}><v>{value}</v></c>'
    text = xml_escape(str(value))
    return f'<c r="{ref}"{style_attr} t="inlineStr"><is><t xml:space="preserve">{text}</t></is></c>'


def row_xml(row_num: int, values: List[Tuple[Any, int, bool]], height: Optional[int] = None) -> str:
    height_attr = f' ht="{height}" customHeight="1"' if height else ""
    cells = "".join(cell_xml(row_num, idx + 1, value, style, numeric) for idx, (value, style, numeric) in enumerate(values))
    return f'<row r="{row_num}"{height_attr}>{cells}</row>'


def styles_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="5">
    <font><sz val="10"/><name val="맑은 고딕"/></font>
    <font><b/><sz val="16"/><color rgb="FFFFFFFF"/><name val="맑은 고딕"/></font>
    <font><b/><sz val="10"/><name val="맑은 고딕"/></font>
    <font><sz val="9"/><color rgb="FF666666"/><name val="맑은 고딕"/></font>
    <font><u/><sz val="9"/><color rgb="FF0563C1"/><name val="맑은 고딕"/></font>
  </fonts>
  <fills count="6">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF0F3B57"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFD9E2F3"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFFCE4D6"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFE7E6E6"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="2">
    <border><left/><right/><top/><bottom/><diagonal/></border>
    <border>
      <left style="thin"><color rgb="FF7F7F7F"/></left>
      <right style="thin"><color rgb="FF7F7F7F"/></right>
      <top style="thin"><color rgb="FF7F7F7F"/></top>
      <bottom style="thin"><color rgb="FF7F7F7F"/></bottom>
      <diagonal/>
    </border>
  </borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="10">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>
    <xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"><alignment horizontal="center" vertical="center"/></xf>
    <xf numFmtId="0" fontId="3" fillId="0" borderId="0" xfId="0"><alignment horizontal="left" vertical="center"/></xf>
    <xf numFmtId="0" fontId="2" fillId="3" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="2" fillId="4" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="2" fillId="5" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1"><alignment horizontal="left" vertical="center" wrapText="1"/></xf>
    <xf numFmtId="3" fontId="0" fillId="0" borderId="1" xfId="0" applyBorder="1" applyNumberFormat="1"><alignment horizontal="right" vertical="center"/></xf>
    <xf numFmtId="0" fontId="4" fillId="0" borderId="1" xfId="0" applyFont="1" applyBorder="1"><alignment horizontal="left" vertical="center" wrapText="1"/></xf>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>"""


def workbook_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="관리대장" sheetId="1" r:id="rId1"/></sheets>
</workbook>"""


def workbook_rels_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>"""


def root_rels_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>
</Relationships>"""


def content_types_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>
  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>
</Types>"""


def doc_props_xml(generated_at: str) -> Tuple[str, str]:
    core = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties"
 xmlns:dc="http://purl.org/dc/elements/1.1/"
 xmlns:dcterms="http://purl.org/dc/terms/"
 xmlns:dcmitype="http://purl.org/dc/dcmitype/"
 xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
  <dc:title>부울경 행사 특별소통 관리대장</dc:title>
  <dc:creator>Codex</dc:creator>
  <dcterms:created xsi:type="dcterms:W3CDTF">{xml_escape(generated_at)}</dcterms:created>
  <dcterms:modified xsi:type="dcterms:W3CDTF">{xml_escape(generated_at)}</dcterms:modified>
</cp:coreProperties>"""
    app = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
 xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">
  <Application>Codex Event Monitor</Application>
</Properties>"""
    return core, app


def sheet_xml(items: List[Dict[str, Any]], generated_at: str) -> str:
    cols = [8, 24, 34, 24, 11, 12, 12, 12, 17, 16, 19, 16, 18, 38, 45, 20, 34, 15]
    col_xml = "".join(f'<col min="{idx}" max="{idx}" width="{width}" customWidth="1"/>' for idx, width in enumerate(cols, 1))
    rows: List[str] = []
    last_col = col_name(len(EVENT_HEADERS))
    rows.append(row_xml(1, [(f"부울경 행사 특별소통 관리대장", 1, False)] + [("", 1, False)] * (len(EVENT_HEADERS) - 1), 28))
    rows.append(row_xml(2, [(f"생성시각: {generated_at} / 신규 감지: {len(items)}건", 2, False)] + [("", 2, False)] * (len(EVENT_HEADERS) - 1), 20))
    header_values = []
    for idx, header in enumerate(EVENT_HEADERS, 1):
        style = 3 if idx <= 7 else 4 if idx <= 11 else 5
        header_values.append((header, style, False))
    rows.append(row_xml(3, header_values, 32))

    for row_idx, item in enumerate(items, 4):
        crowd_numeric = item.get("crowd_value")
        crowd_value: Any = crowd_numeric if isinstance(crowd_numeric, int) else item.get("crowd", "(자료 없음)")
        values = [
            (row_idx - 3, 6, True),
            (item.get("period", ""), 7, False),
            (item.get("title", ""), 7, False),
            (item.get("place", ""), 7, False),
            (item.get("type", ""), 6, False),
            (item.get("start_short", ""), 6, False),
            (item.get("end_short", ""), 6, False),
            (item.get("grade", ""), 6, False),
            (item.get("feature", ""), 6, False),
            (item.get("network", ""), 6, False),
            (crowd_value, 8 if isinstance(crowd_value, int) else 6, isinstance(crowd_value, int)),
            (item.get("source_name", ""), 6, False),
            (item.get("published", ""), 7, False),
            (item.get("link", ""), 9, False),
            (item.get("summary", ""), 7, False),
            (item.get("collected_at", ""), 6, False),
            (item.get("extraction_method", ""), 7, False),
            (item.get("status", ""), 6, False),
        ]
        rows.append(row_xml(row_idx, values, 58))

    dimension = f"A1:{last_col}{max(4, len(items) + 3)}"
    merge_xml = f'<mergeCells count="2"><mergeCell ref="A1:{last_col}1"/><mergeCell ref="A2:{last_col}2"/></mergeCells>'
    auto_filter = f'<autoFilter ref="A3:{last_col}{max(4, len(items) + 3)}"/>'
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <dimension ref="{dimension}"/>
  <sheetViews><sheetView workbookViewId="0"><pane ySplit="3" topLeftCell="A4" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>
  <sheetFormatPr defaultRowHeight="16"/>
  <cols>{col_xml}</cols>
  <sheetData>{''.join(rows)}</sheetData>
  {merge_xml}
  {auto_filter}
  <pageMargins left="0.3" right="0.3" top="0.5" bottom="0.5" header="0.3" footer="0.3"/>
</worksheet>"""


def write_xlsx(items: List[Dict[str, Any]], output_path: Path, generated_at: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "관리대장"

    title = "부울경 행사 특별소통 관리대장"
    subtitle = f"생성시각: {generated_at} / 신규 감지: {len(items)}건"

    ws.append([title])
    ws.append([subtitle])
    ws.append(EVENT_HEADERS)

    # Title merge
    last_col = len(EVENT_HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)

    # Styles
    title_fill = PatternFill("solid", fgColor="0F3B57")
    header_fill_left = PatternFill("solid", fgColor="D9E2F3")
    header_fill_mid = PatternFill("solid", fgColor="FCE4D6")
    header_fill_right = PatternFill("solid", fgColor="E7E6E6")

    white_font = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="맑은 고딕", size=10, bold=True)
    normal_font = Font(name="맑은 고딕", size=10)
    link_font = Font(name="맑은 고딕", size=9, color="0563C1", underline="single")

    thin = Side(style="thin", color="7F7F7F")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title rows
    ws["A1"].font = white_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center

    ws["A2"].font = Font(name="맑은 고딕", size=9, color="666666")
    ws["A2"].alignment = left

    # Header row
    for col_idx, header in enumerate(EVENT_HEADERS, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = center
        cell.border = border

        if col_idx <= 7:
            cell.fill = header_fill_left
        elif col_idx <= 11:
            cell.fill = header_fill_mid
        else:
            cell.fill = header_fill_right

    # Data rows
    for row_idx, item in enumerate(items, start=4):
        crowd_numeric = item.get("crowd_value")
        crowd_value = crowd_numeric if isinstance(crowd_numeric, int) else item.get("crowd", "(자료 없음)")

        row_values = [
            row_idx - 3,
            item.get("period", ""),
            item.get("title", ""),
            item.get("place", ""),
            item.get("type", ""),
            item.get("start_short", ""),
            item.get("end_short", ""),
            item.get("grade", ""),
            item.get("feature", ""),
            item.get("network", ""),
            crowd_value,
            item.get("source_name", ""),
            item.get("published", ""),
            item.get("link", ""),
            item.get("summary", ""),
            item.get("collected_at", ""),
            item.get("extraction_method", ""),
            item.get("status", ""),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = normal_font
            cell.border = border

            if col_idx in [3, 4, 14, 15, 17]:
                cell.alignment = left
            else:
                cell.alignment = center

            # 원문링크 컬럼
            if col_idx == 14 and value:
                cell.hyperlink = str(value)
                cell.value = "원문"
                cell.font = link_font
                cell.alignment = center

        ws.row_dimensions[row_idx].height = 58

    # Column widths
    widths = [8, 24, 34, 28, 11, 12, 12, 12, 17, 16, 19, 16, 18, 18, 45, 20, 36, 15]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 32

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(last_col)}{max(4, len(items) + 3)}"

    wb.save(output_path)


def build_html_report(items: List[Dict[str, Any]], generated_at: str) -> str:
    rows = []
    for item in items:
        rows.append(
            f"""
            <tr>
              <td>{escape(item.get('grade', ''))}</td>
              <td><strong>{escape(item.get('title', ''))}</strong><br><span>{escape(item.get('summary', ''))}</span></td>
              <td>{escape(item.get('period', ''))}</td>
              <td>{escape(item.get('place', ''))}</td>
              <td>{escape(item.get('type', ''))}</td>
              <td>{escape(str(item.get('crowd', '')))}</td>
              <td>{escape(item.get('extraction_method', ''))}</td>
              <td><a href="{escape(item.get('link', ''))}">원문</a></td>
            </tr>
            """
        )
    body = "".join(rows) if rows else '<tr><td colspan="8">새 게시물이 없습니다.</td></tr>'
    return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <title>부울경 행사 특별소통 알림</title>
  <style>
    body {{ font-family: "Malgun Gothic", "Segoe UI", sans-serif; margin: 24px; color: #17212b; }}
    h1 {{ margin: 0 0 6px; font-size: 22px; }}
    .meta {{ color: #5f6f7a; margin-bottom: 16px; }}
    table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
    th {{ background: #0f3b57; color: white; padding: 9px; border: 1px solid #9aa7b0; }}
    td {{ padding: 9px; border: 1px solid #c7d0d8; vertical-align: top; }}
    tr:nth-child(even) td {{ background: #f7fafc; }}
    span {{ color: #495762; }}
  </style>
</head>
<body>
  <h1>부울경 행사 특별소통 알림</h1>
  <div class="meta">생성시각: {escape(generated_at)} / 신규 감지: {len(items)}건</div>
  <table>
    <thead>
      <tr>
        <th>등급</th><th>Event명/요약</th><th>기간</th><th>장소</th><th>Type</th><th>예상운집</th><th>추출방식</th><th>링크</th>
      </tr>
    </thead>
    <tbody>{body}</tbody>
  </table>
</body>
</html>"""


def send_outlook_mail(config: Dict[str, Any], items: List[Dict[str, Any]], html_body: str, attachment: Path, display: bool) -> None:
    mail_cfg = config.get("mail", {})
    to = clean_text(mail_cfg.get("to", ""))
    cc = clean_text(mail_cfg.get("cc", ""))
    if not to:
        print("메일 수신자가 비어 있어 Outlook 메일은 만들지 않았습니다.")
        return
    if win32com is None:
        print("win32com.client를 사용할 수 없어 Outlook 메일은 만들지 않았습니다.")
        return
    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = to
    if cc:
        mail.CC = cc
    prefix = clean_text(mail_cfg.get("subject_prefix", "[부울경 행사 특별소통 알림]"))
    mail.Subject = f"{prefix} 신규 {len(items)}건"
    mail.HTMLBody = html_body
    if attachment.exists():
        mail.Attachments.Add(str(attachment))
    mode = "display" if display else clean_text(mail_cfg.get("mode", "display")).lower()
    if mode == "send":
        mail.Send()
    else:
        mail.Display()

def send_smtp_mail(config: Dict[str, Any], items: List[Dict[str, Any]], html_body: str, attachments: List[Path]) -> None:
    mail_cfg = config.get("mail", {})

    mail_user = os.getenv("MAIL_USERNAME", "").strip()
    mail_password = os.getenv("MAIL_PASSWORD", "").strip()
    mail_to = os.getenv("MAIL_TO", "").strip() or clean_text(mail_cfg.get("to", ""))
    mail_cc = os.getenv("MAIL_CC", "").strip() or clean_text(mail_cfg.get("cc", ""))

    mail_from = os.getenv("MAIL_FROM", "").strip() or mail_user
    smtp_host = os.getenv("SMTP_HOST", "").strip() or "smtp.gmail.com"
    smtp_port = int(os.getenv("SMTP_PORT", "").strip() or "587")
    smtp_use_tls = (os.getenv("SMTP_USE_TLS", "").strip().lower() or "true") in {"true", "1", "yes", "y"}

    subject_prefix = (
        os.getenv("MAIL_SUBJECT_PREFIX", "").strip()
        or clean_text(mail_cfg.get("subject_prefix", "[부울경 행사 특별소통 알림]"))
    )

    if not mail_user or not mail_password:
        print("MAIL_USERNAME 또는 MAIL_PASSWORD가 없어 SMTP 메일을 보내지 않았습니다.")
        return

    if not mail_to:
        print("MAIL_TO가 없어 SMTP 메일을 보내지 않았습니다.")
        return

    generated_at = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")
    subject = f"{subject_prefix} 신규 {len(items)}건 ({generated_at})"

    recipients = [addr.strip() for addr in mail_to.split(",") if addr.strip()]
    cc_recipients = [addr.strip() for addr in mail_cc.split(",") if addr.strip()]
    all_recipients = recipients + cc_recipients

    msg = MIMEMultipart()
    msg["From"] = mail_from
    msg["To"] = ", ".join(recipients)
    if cc_recipients:
        msg["Cc"] = ", ".join(cc_recipients)
    msg["Subject"] = subject

    msg.attach(MIMEText(html_body, "html", "utf-8"))

    for attachment in attachments:
        if attachment and attachment.exists():
            with attachment.open("rb") as f:
                part = MIMEApplication(f.read(), Name=attachment.name)
            part["Content-Disposition"] = f'attachment; filename="{attachment.name}"'
            msg.attach(part)

    with smtplib.SMTP(smtp_host, smtp_port) as server:
        if smtp_use_tls:
            server.starttls()
        server.login(mail_user, mail_password)
        server.sendmail(mail_from, all_recipients, msg.as_string())

    print("SMTP 이메일 발송 완료")



def open_output(path: Path) -> None:
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore[attr-defined]
    except Exception as exc:
        print(f"자동 열기 실패: {exc}")


def main() -> int:
    parser = argparse.ArgumentParser(description="부울경 행사 모니터링 통합본")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="설정 JSON 경로")
    parser.add_argument("--state", default=str(DEFAULT_STATE), help="중복 확인 상태 JSON 경로")
    parser.add_argument("--dry-run", action="store_true", help="메일 발송 없이 수집/엑셀 생성만 실행")
    parser.add_argument("--display", action="store_true", help="Outlook 메일 창만 열기")
    parser.add_argument("--include-seen", action="store_true", help="이미 본 항목도 결과에 포함")
    parser.add_argument("--reset-state", action="store_true", help="중복 기록 초기화 후 실행")
    parser.add_argument("--no-open", action="store_true", help="결과 파일 자동 열기 생략")
    args = parser.parse_args()

    ensure_runtime_files()
    config_path = Path(args.config)
    config = read_json(config_path, DEFAULT_CONFIG_DATA)
    state_path = Path(args.state)
    if args.reset_state and state_path.exists():
        state_path.unlink()
    state = read_json(state_path, {"seen_ids": {}})

    all_items, errors = collect_items(config)
    new_items = dedupe_new_items(all_items, state, args.include_seen)
    new_items = apply_limits(new_items, config)
    write_json(state_path, state)

    generated_at = datetime.now(KST).strftime("%Y-%m-%d %H:%M:%S")
    file_tag = datetime.now(KST).strftime("%Y-%m-%d_%H%M")
    run_id = datetime.now(KST).strftime("%Y%m%d_%H%M%S")
    xlsx_path = REPORT_DIR / f"{file_tag}_부울경행사_신규관리대장.xlsx"
    all_xlsx_path = REPORT_DIR / f"{file_tag}_부울경행사_전체관리대장.xlsx"
    html_path = REPORT_DIR / f"{file_tag}_부울경행사_요약.html"
    write_xlsx(new_items, xlsx_path, generated_at)
    write_xlsx(sort_items(all_items), all_xlsx_path, generated_at)
    write_xlsx(new_items, xlsx_path, generated_at)
    html_body = build_html_report(new_items, generated_at)
    html_path.write_text(html_body, encoding="utf-8")
    write_json(DATA_DIR / "results.json", [public_item(item) for item in sort_items(all_items)])
    write_json(DATA_DIR / "new_items.json", [public_item(item) for item in new_items])
    append_history(DATA_DIR / "history.jsonl", run_id, new_items)
    write_json(
        DEFAULT_RUN_SUMMARY,
        {
            "run_id": run_id,
            "generated_at": generated_at,
            "candidate_count": len(all_items),
            "new_count": len(new_items),
            "xlsx_path": str(xlsx_path),
            "html_path": str(html_path),
            "article_detail_stats": ARTICLE_DETAIL_STATS,
            "ai_extraction_stats": AI_EXTRACTION_STATS,
            "geocoder_stats": GEOCODER_STATS,
            "errors": errors,
        },
    )

    print(f"전체 감지 후보: {len(all_items)}건")
    print(f"신규/출력 대상: {len(new_items)}건")
    print(f"엑셀 저장: {xlsx_path}")
    print(f"요약 저장: {html_path}")
    print(f"기사 본문 조회 통계: {ARTICLE_DETAIL_STATS}")
    print(f"AI 추출 통계: {AI_EXTRACTION_STATS}")
    print(f"장소 주소 변환 통계: {GEOCODER_STATS}")
    if errors:
        print("수집 오류:")
        for error in errors:
            print(f"- {error}")

    if not args.dry_run and new_items:
        send_smtp_mail(config, new_items, html_body, [xlsx_path, all_xlsx_path])
    elif args.dry_run:
        print("--dry-run 이라서 메일은 만들지 않았습니다.")
    else:
        print("새 항목이 없어 메일은 만들지 않았습니다.")

    if config.get("output", {}).get("open_after_run", True) and not args.no_open:
        open_output(xlsx_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
